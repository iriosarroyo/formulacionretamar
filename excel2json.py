from openpyxl import load_workbook

filename = "C:/Users/Inigo/Documents/SpeedRuns/formulacion/compuestos_inorganicos_hacer_examenes.xlsm"
wb = load_workbook(filename)

sheets = wb.sheetnames
json_data = {}

for sheet in sheets:
    json_data[sheet] = []
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=True):
        if not row[1] or not row[2]:
            continue
        elems = [row[1], row[2]]
        json_data[sheet].append(elems)

print(
    "export const formInorganica =",
    json_data,
    file=open("./public/formulacionInorganica.js", "w", encoding="utf-8"),
)
